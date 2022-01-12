from openpyxl import load_workbook
import mysql.connector

# Excel
workbook = load_workbook('imported_02.xlsx') #สร้างตัวแปรไว้อ่านไฟล์excel
sheet = workbook.active

#เชื่อมต่อDatabase
db = mysql.connector.connect(
    host='localhost',
    port=3306,
    user='root',
    password='Title@20572',
    database='title_database'

)
cursor = db.cursor()

#ให้ทำการโหลดข้อมูลประเภทสินค้าทั้งหมด
sql_select_catagories = '''
    select *
    from catagories

'''

cursor.execute(sql_select_catagories)
catagories = cursor.fetchall()

#เปรียบเทียบข้อมูลประเภทสินค้าอันไหนยังไม่มีให้เพิ่มลงไป
catagories_values = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    is_new = True
    category = row[3]

    for c in catagories:
        if category == c[1]:
            is_new = False
            break

    if is_new:
        catagories_values.append((category,))

if len(catagories_values) > 0:
    sql_insert_catagories = '''
        insert into catagories (title)
        values (%s)
    
    '''
    cursor.executemany(sql_insert_catagories,catagories_values)
    db.commit()
    print('เพิ่มประเภทสินค้าจำนวน ' + str(cursor.rowcount))

else:
    print('ไม่มีสินค้ามาเพิ่ม')
#โหลดข้อมูลประเภทสินค้าทั้งหมดอีกครังนึง
cursor.execute(sql_select_catagories)
catagories = cursor.fetchall()


#เชื่อมต่อ categorie id สินค้าใหม่ของเรา เเล้วเพิ่มลงไป
products_value = []
for row in sheet.iter_rows(min_row=2,values_only=True):
    category_title = row[3]
    category_id = 'Null'

    for c in catagories:
        if category_title == c[1]:
            category_id = c[0]
            break


    product = (row[0],row[1],row[2],category_id)
    print(product)

    products_value.append(product)

sql_insert_product = '''
    insert into products (title ,price, is_necessary,catagory_id)
    value(%s,%s,%s,%s)

'''
cursor.executemany(sql_insert_product,products_value)
db.commit()
print('เพิ่มสินค้าทั้งหมด' + str(cursor.rowcount))


cursor.close()
db.close()